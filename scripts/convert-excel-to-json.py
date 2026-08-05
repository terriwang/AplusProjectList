#!/usr/bin/env python3
"""
convert-excel-to-json.py

Converts one or more HDA (Housing Delivery Authority) Excel workbooks
(a main database export, plus optional addendum exports) into a single
structured JSON file used by the HDA Intelligence Database website.

This script does not alter, correct, guess, or enrich any project data.
It only reshapes workbook rows into JSON, and performs light,
non-destructive normalisation needed for the interface to work
(trimming whitespace, converting dates to ISO format, splitting
long-form text fields into bullet points, and resolving duplicate or
renamed column headers).

Usage:
    python convert-excel-to-json.py path/to/HDA_Database.xlsx
    python convert-excel-to-json.py path/to/HDA_Database.xlsx path/to/Addendum.xlsx [more files...]

When multiple files are given, every row from every file is included in
the output (nothing is merged or de-duplicated automatically), in the
order the files were listed. If the same EOI Reference Number appears in
more than one file, a warning is printed so you can check whether that's
expected (e.g. a genuine addendum correction) rather than an accidental
duplicate — both rows are still kept, since the script never discards or
guesses about source data.

Output:
    data/hda-projects.json (relative to the project root, i.e. one
    directory above this script)
"""

import sys
import json
import re
import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# Columns that must resolve to a single 'status' value taken from this
# fixed set. Only the LAST column matching this name (case-insensitive)
# is used, per the brief's instruction to use the "final valid" status
# column when duplicates exist.
VALID_STATUSES = {"Recommended", "Deferred", "Not Recommended"}

# Canonical -> list of alternate header spellings we should still
# recognise if the source workbook changes slightly. The FIRST canonical
# name whose alternates match a header wins; if a workbook has genuinely
# duplicate headers (e.g. two columns both literally called "Status"),
# the later one in column order overrides the earlier one, and a warning
# is printed.
HEADER_ALIASES = {
    "date": ["date"],
    "internal_ref": ["internal reference number"],
    "report_item_no": ["report item no.", "report item no"],
    "eoi_ref": ["eoi reference number"],
    "status": ["status"],
    "address": ["address"],
    "suburb": ["suburb"],
    "lga": ["lga"],
    "developer": ["developer"],
    "planning_consultant": ["planning consultant"],
    "architect": ["architect"],
    "applicant_raw": ["applicant (raw)", "applicant"],
    "project_type": ["project type"],
    "storeys": ["storeys"],
    "dwellings": ["no. of dwellings", "no of dwellings", "number of dwellings"],
    "concurrent_rezoning": ["concurrent rezoning"],
    "affordable_housing_pct": ["affordable housing %", "affordable housing percent"],
    "affordable_housing_tenure": ["affordable housing tenure"],
    "reason_full": ["reason (full)", "reason full"],
    "script": ["script", "minister / applicant advice", "minister/applicant advice", "minister applicant advice"],
    "criteria": ["criteria", "rules"],
    "notes": ["notes"],
}


def normalise_header(raw):
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw)).strip().lower()


def clean_text(value):
    """Trim whitespace, normalise line breaks, preserve content verbatim."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\r\n", "\n").replace("\r", "\n").strip()
        return text if text != "" else None
    return value


def to_iso_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    # Fall back: try to parse common string date formats
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return text  # leave as-is if unrecognised, rather than guessing


def to_number_or_none(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return None  # non-numeric (e.g. a range like "4-25") -> keep raw string elsewhere


def storeys_sort_key(value):
    """Extract the first integer in a storeys value (which may be a
    single number, a range like '4-25', or blank) for numeric sorting.
    Returns None if no number can be found, so blanks sort last."""
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


def split_bullets(text):
    """Split a long-form field that uses '•' (top level) and 'o ' (sub
    level, following a top-level bullet) markers into a structured list
    of {level, text} items. Content is preserved verbatim; only the
    markers are used to determine structure and are stripped from the
    item text itself."""
    if not text:
        return []

    # Split on top-level bullets first, keeping sub-bullets attached.
    parts = re.split(r"(?<!\S)•\s*", text)
    items = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # A top-level bullet may itself contain one or more "o " sub-points.
        sub_parts = re.split(r"(?<=\.)\s+o\s+|(?<=:)\s+o\s+|^\s*o\s+", part)
        # The first sub_parts element is the top-level text itself.
        top = sub_parts[0].strip()
        if top:
            items.append({"level": 0, "text": top})
        for sub in sub_parts[1:]:
            sub = sub.strip()
            if sub:
                items.append({"level": 1, "text": sub})
    return items


def normalise_concurrent_rezoning(raw):
    """The source workbooks have used two different vocabularies for this
    field over time ('Concurrent Rezoning' / 'None', and 'Yes' / 'No' /
    'Unknown'). Normalise both into one of: 'yes', 'no', 'unknown', or
    None (blank). Anything else is preserved as 'other' so the raw text
    is never silently dropped."""
    if raw is None:
        return None
    text = raw.strip().lower()
    if text == "":
        return None
    if text in ("yes", "concurrent rezoning"):
        return "yes"
    if text in ("no", "none"):
        return "no"
    if text == "unknown":
        return "unknown"
    return "other"


def parse_criteria(text):
    """Parse a semicolon-separated 'Objective N: ...; Criteria N.N: ...'
    field into a list of {"label": ..., "description": ...} items,
    preserving the source wording. Returns [] if the field is empty."""
    if not text:
        return []
    items = []
    for chunk in text.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        m = re.match(r"^((?:Objective|Criteria|Criterion)\s*[\d.]+)\s*[:\-–]\s*(.+)$", chunk, re.IGNORECASE)
        if m:
            items.append({"label": m.group(1).strip(), "description": m.group(2).strip()})
        else:
            items.append({"label": None, "description": chunk})
    return items


def resolve_headers(header_row):
    """Map each column index to a canonical field key. Handles duplicate
    headers by keeping the LAST occurrence of a given canonical field
    and warning about it."""
    lookup = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            lookup[alias] = canonical

    col_to_field = {}
    seen_canonical_cols = {}
    warnings = []

    for idx, raw_header in enumerate(header_row):
        norm = normalise_header(raw_header)
        canonical = lookup.get(norm)
        if canonical is None:
            warnings.append(f"Column {idx} ('{raw_header}') did not match a known field and will be ignored.")
            continue
        if canonical in seen_canonical_cols:
            warnings.append(
                f"Duplicate column for field '{canonical}': column {seen_canonical_cols[canonical]} "
                f"('{header_row[seen_canonical_cols[canonical]]}') and column {idx} ('{raw_header}'). "
                f"Using the later column (index {idx}) as the final value."
            )
        seen_canonical_cols[canonical] = idx
        col_to_field[idx] = canonical

    # If a field appeared more than once, only the LAST column index should
    # actually be used to populate it; earlier occurrences of that same
    # canonical field are dropped from col_to_field so they don't overwrite
    # the final value when iterating left-to-right... but since we iterate
    # left to right, later overwrites earlier naturally if we just keep
    # writing into the same dict key during row parsing. No extra action
    # needed here.
    return col_to_field, warnings


def find_database_sheet(workbook):
    """Pick the worksheet that looks like the HDA project database."""
    candidates = []
    for name in workbook.sheetnames:
        if "hda" in name.lower() or "database" in name.lower():
            candidates.append(name)
    if candidates:
        return candidates[0]
    # Fall back to the sheet with the most rows.
    return max(workbook.sheetnames, key=lambda n: workbook[n].max_row)


def process_workbook(input_path, file_index):
    """Read one workbook and return (projects, stats) where stats holds
    counts/warnings for the summary printed at the end."""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    sheet_name = find_database_sheet(wb)
    ws = wb[sheet_name]
    print(f"\n[{input_path.name}] Using worksheet: '{sheet_name}'")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"WARNING: [{input_path.name}] Worksheet is empty, skipping this file.")
        return [], {"skipped_empty": 0, "status_mismatches": []}

    header_row = rows[0]
    data_rows = rows[1:]

    col_to_field, warnings = resolve_headers(header_row)
    for w in warnings:
        print(f"WARNING: [{input_path.name}] {w}")

    projects = []
    skipped_empty = 0
    status_mismatches = []

    for row_num, row in enumerate(data_rows, start=2):
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            skipped_empty += 1
            continue

        record = {}
        for idx, field in col_to_field.items():
            value = row[idx] if idx < len(row) else None
            record[field] = value

        status_raw = clean_text(record.get("status"))
        if status_raw not in VALID_STATUSES:
            status_mismatches.append((row_num, status_raw))

        storeys_raw = clean_text(record.get("storeys"))
        reason_full = clean_text(record.get("reason_full"))
        script_text = clean_text(record.get("script"))
        criteria_raw = clean_text(record.get("criteria"))
        concurrent_raw = clean_text(record.get("concurrent_rezoning"))
        aff_pct_raw = clean_text(record.get("affordable_housing_pct"))
        aff_tenure_raw = clean_text(record.get("affordable_housing_tenure"))
        concurrent_status = normalise_concurrent_rezoning(concurrent_raw)

        project = {
            "id": f"src{file_index}-row-{row_num}",
            "sourceFile": input_path.name,
            "date": to_iso_date(record.get("date")),
            "internalReferenceNumber": clean_text(record.get("internal_ref")),
            "reportItemNo": clean_text(record.get("report_item_no")),
            "eoiReferenceNumber": clean_text(record.get("eoi_ref")),
            "status": status_raw,
            "address": clean_text(record.get("address")),
            "suburb": clean_text(record.get("suburb")),
            "lga": clean_text(record.get("lga")),
            "developer": clean_text(record.get("developer")),
            "planningConsultant": clean_text(record.get("planning_consultant")),
            "architect": clean_text(record.get("architect")),
            "applicantRaw": clean_text(record.get("applicant_raw")),
            "projectType": clean_text(record.get("project_type")),
            "storeys": storeys_raw,
            "storeysSort": storeys_sort_key(storeys_raw),
            "dwellings": to_number_or_none(record.get("dwellings")),
            "concurrentRezoning": concurrent_status == "yes",
            "concurrentRezoningStatus": concurrent_status,
            "concurrentRezoningRaw": concurrent_raw,
            "affordableHousingPct": None if (aff_pct_raw and aff_pct_raw.strip().lower() == "none") else aff_pct_raw,
            "affordableHousingTenure": None if (aff_tenure_raw and aff_tenure_raw.strip().lower() == "none") else aff_tenure_raw,
            "reasonFull": reason_full,
            "reasonFullItems": split_bullets(reason_full),
            "script": script_text,
            "criteriaRaw": criteria_raw,
            "criteriaItems": parse_criteria(criteria_raw),
            "notes": clean_text(record.get("notes")),
        }
        projects.append(project)

    if status_mismatches:
        print(f"WARNING: [{input_path.name}] {len(status_mismatches)} row(s) have a status "
              f"value outside {sorted(VALID_STATUSES)}:")
        for row_num, val in status_mismatches[:20]:
            print(f"   - row {row_num}: {val!r}")

    print(f"[{input_path.name}] Skipped {skipped_empty} fully empty row(s). "
          f"Exported {len(projects)} project row(s).")

    return projects, {"skipped_empty": skipped_empty, "status_mismatches": status_mismatches}


def main():
    if len(sys.argv) < 2:
        print("Usage: python convert-excel-to-json.py path/to/HDA_Database.xlsx [path/to/Addendum.xlsx ...]")
        sys.exit(1)

    input_paths = [Path(p) for p in sys.argv[1:]]
    for p in input_paths:
        if not p.exists():
            print(f"ERROR: File not found: {p}")
            sys.exit(1)

    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent
    output_path = project_root / "data" / "hda-projects.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    all_projects = []
    source_descriptors = []
    for file_index, input_path in enumerate(input_paths, start=1):
        projects, stats = process_workbook(input_path, file_index)
        all_projects.extend(projects)
        source_descriptors.append({"file": input_path.name, "projectCount": len(projects)})

    # Warn (but never drop rows) if the same EOI Reference Number shows up
    # in more than one input file — could be a genuine addendum correction
    # or an accidental duplicate; either way it's worth a human glance.
    eoi_sources = {}
    for p in all_projects:
        eoi = p.get("eoiReferenceNumber")
        if not eoi:
            continue
        eoi_sources.setdefault(eoi, []).append(p["sourceFile"])
    cross_file_duplicates = {eoi: files for eoi, files in eoi_sources.items() if len(set(files)) > 1}
    if cross_file_duplicates:
        print(f"\nWARNING: {len(cross_file_duplicates)} EOI Reference Number(s) appear in more than one "
              f"input file (all rows are kept — check whether this is expected):")
        for eoi, files in list(cross_file_duplicates.items())[:20]:
            print(f"   - EOI {eoi}: {', '.join(sorted(set(files)))}")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "sourceFiles": source_descriptors,
                "generatedAt": datetime.datetime.now().isoformat(timespec="seconds"),
                "projectCount": len(all_projects),
                "projects": all_projects,
            },
            f,
            indent=2,
            ensure_ascii=False,
        )

    print(f"\nExported {len(all_projects)} total project(s) from {len(input_paths)} file(s) to {output_path}")

    status_counts = {}
    for p in all_projects:
        status_counts[p["status"]] = status_counts.get(p["status"], 0) + 1
    print("Status breakdown:", status_counts)


if __name__ == "__main__":
    main()
